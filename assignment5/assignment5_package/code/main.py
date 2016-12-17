"""
.. module:: main
   :platform: Unix, Windows
   :synopsis: Load an excel file with data to calculate the values of
   linear regression parameters (ßø y ß¡) and correlation parameters (r y r²).

.. moduleauthor:: Jaime Jimenez


"""
import os
from LinkedList import LinkedList
from ProbeMethod import LinearRegression, ProbeMethodVals, CorrelationFactor
from openpyxl import load_workbook, Workbook

def load_data_excel(file):
    """Load the data file in an array. The cells are given by cell_range.
    Stop looping in file when find an empty cell
    Args: 
       file: file path to load.
    Returns: 
       Array.  The array populated with the data gotten from file 
    >>> print load_data_excel(self, file)

    """
    cell_range = 'B1:E500'
    wb = load_workbook(filename=file)
    ws = wb.active
    arr = []
    head = True
    for row in ws.iter_rows(cell_range):
        r = []
        for cell in row:
            if cell.value is None or cell.value == '':
                break
            if head:
                r.append(cell.value)
            else:
                r.append(float(cell.value))
        else:
            if head:
                head = False
            arr.append(r)
            continue
        break
    return arr

def pop_linked_list(arr, hcols, elems_x_node):
    """Populate the linkedlist with the values in arr, taking the columns cols
    Args: 
       arr: Array multidimensional with the values to populate the linkedlist.
       cols: Array with the values of the columns to take the values from arr.
       elems_x_node : number or elements per node. 
    Returns: 
       LinkedList. Linked list populated 
    >>> print pop_linked_list(self, arr, cols)

    """
    linked_list = LinkedList()
    head = True
    cols = []
    for row in arr:
        node_value = []
        if head:
            head = False
            i = 0
            while i < len(row):
                for hcol in hcols:
                    if hcol == row[i]:
                        cols.append(i)
                i += 1
        else:
            for col in cols:
                node_value.append(row[col])
            if len(node_value) == elems_x_node:
                linked_list.add(node_value)
            else:
                raise ValueError("The number of elements per node must be "+
                    str(elems_x_node)+". "+str(len(node_value))+" getted: "+str(node_value))
    return linked_list
            
def print_cells(arr, file, sheet):
    """Print the array arr in sheet of excel file
    Args: 
       arr: Array multidimensional with the values to print in file.
       file: Excel file to print the result.
       sheet: Sheet in excel file to print.
    Returns: 
       LinkedList. Linked list populated 
    >>> print pop_linked_list(self, arr, cols)

    """
    wb = load_workbook(filename=file)
    ws = wb.create_sheet(title=sheet)
    j = 1
    for row in arr:
        i = 1
        for col in row:
            _ = ws.cell(column = i , row = j, value="%s" % col)
            i += 1
        j += 1
    wb.save(filename=file)
file_name = "values.xlsx"
#Start program
path = input("\n\nCual es el archivo excel que desea evaluar? \
    Para evaluar values.xlsx ubicado en la misma carpeta del programa \
    presione enter.\n\n")
path = path.strip()
if path == "":
    path = os.path.join(os.getcwd(),file_name)
if os.path.isdir(path):
    path = os.path.join(path,file_name)
if not os.path.isfile(path):
    raise ValueError("Ha ocurrido un error con el archivo seleccionado.")

estimated_proxy_size = int(input("Cual es el proxy size estimado?.\n\n\n"))

arr_data = load_data_excel(file = path);
print("")
print("")
print(arr_data)
# Cols to get from arr_data and calculate the formulas
# arr_cols is compared with the excel header to get the values
arr_cols = [['ESTIMATED_PROXY_SIZE','ACTUAL_ADDED_MODIFIED_SIZE'],
['ESTIMATED_PROXY_SIZE','ACTUAL_DEV_HOURS'],
['PLAN_ADDED_MODIFIED_SIZE','ACTUAL_ADDED_MODIFIED_SIZE'],
['PLAN_ADDED_MODIFIED_SIZE','ACTUAL_DEV_HOURS']]
arr_result = []
i = 0
arr_result.append(["ßø", "ß¡", "r", "r²", "P"])
print("")
print("")
print(arr_result[i])
for cols in arr_cols:
    llist_data = pop_linked_list(arr_data, cols, 2)
    probe_vls = ProbeMethodVals(llist_data) # vals to calculate the formulas
    lr = LinearRegression(probe_vals = probe_vls, E = estimated_proxy_size)
    cr = CorrelationFactor(probe_vals = probe_vls)
    arr_result.append([round(lr.get_beta_0(), 4), round(lr.get_beta_1(), 4),
        round(cr.get_r(), 4), round(cr.get_r_squared(), 4), round(lr.get_P(), 4)])
    i += 1
    print(arr_result[i])

    
print_cells(arr_result, path, "restult")

print("\n\nPuede revisar los resultados en la pestaña results del \
    Archivo "+path)

print("[FIN]")

