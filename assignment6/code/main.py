"""
.. module:: main
   :platform: Unix, Windows
   :synopsis: Load an excel file with data to calculate the relative
   sizes.

.. moduleauthor:: Jaime Jimenez


"""
import os
import sys
from openpyxl import load_workbook, Workbook
from RelativeSize import RelativeSize

def get_path(file_name):
    """
    Get the file path from user input or from default file and current path
    Args: 
       file_name: default file .
    Returns: 
       String. The full path of excel file
    >>> print get_path(file)

    """
    path = input("\n\nCual es el archivo excel que desea evaluar? \
        Para evaluar values.xlsx ubicado en la misma carpeta del programa \
        presione enter.\n")
    path = path.strip()
    if path == "":
        path = os.path.join(os.getcwd(),file_name)
        print(path)
    if os.path.isdir(path):
        path = os.path.join(path,file_name)
    if not os.path.isfile(path):
        raise ValueError("Ha ocurrido un error con el archivo seleccionado.")
    return path

def load_excel_data(headers, file):
    """
    Load the excel data and search the headers. Returns an array with the
    values of the cells in the same cols where the headers were found.
    Args:
       headers: Array with the value of the Headers. To be compare with every 
       column in the excel file searching for the matches
       file: Full file path
    Returns:
       Array: The array with the cells values in the columns that matched with
       with the header.

    """
    wb = load_workbook(filename=file)
    ws = wb.active
    arr = []
    cols = []
    row = 2
    for col in range(1,100): # Search in row 1 for the header columns
        if len(cols) > 2: # Just accepted one or two columns to evaluate
            break
        val = ws.cell(row=1, column=col).value
        if val == headers[0] or val == headers[1]:
            cols.append(col)
    if len(cols) < 1:
        raise IOError("El archivo no contiene las columnas requeridas")
    while True:
        r = []
        for col in cols:
            val = ws.cell(row=row, column=col).value
            if val == "" or val is None:
                break
            r.append(float(val))
        else:
            arr.append(r)
            row += 1
            continue
        break
    return arr

def print_relarive_sizes(rs):
    print("")
    print("Relative Sizes:")
    print("XS:"+str(rs.XS))
    print("S:"+str(rs.S))
    print("M:"+str(rs.M))
    print("L:"+str(rs.L))
    print("XL:"+str(rs.XL))

def write_relarive_sizes(rs, tab, file):
    wb = load_workbook(filename=file)
    if tab in wb.get_sheet_names():
        wb.remove_sheet(wb.get_sheet_by_name(tab))
    ws = wb.create_sheet(title=tab)

    _ = ws.cell(column = 1 , row = 1, value="XS")
    _ = ws.cell(column = 2 , row = 1, value="S")
    _ = ws.cell(column = 3 , row = 1, value="M")
    _ = ws.cell(column = 4 , row = 1, value="L")
    _ = ws.cell(column = 5 , row = 1, value="XL")
    _ = ws.cell(column = 1 , row = 2, value="%s" % rs.XS)
    _ = ws.cell(column = 2 , row = 2, value="%s" % rs.S)
    _ = ws.cell(column = 3 , row = 2, value="%s" % rs.M)
    _ = ws.cell(column = 4 , row = 2, value="%s" % rs.L)
    _ = ws.cell(column = 5 , row = 2, value="%s" % rs.XL)
    wb.save(filename=file)

def main(argv=None):
    """Start point. 
    Args: 
       argv: Arguments.
    >>> print main(argv=None)
    
    """
    __file_name = "values.xlsx" # Default file name to calculate the relative sizes
    __headers = ['UNIT_MESSURE','PARTS'] # Header to search in excel file
    path = get_path(__file_name)
    arr = load_excel_data(__headers, path)
    rs = RelativeSize(arr)
    print_relarive_sizes(rs)
    write_relarive_sizes(rs, tab="restult", file =path)
    return 2

if __name__ == "__main__":
    sys.exit(main(sys.argv))
